import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# マスタデータサンプル
master_data = [
    ['完成品001', 'ノートパソコンセット', '部品001', '部品001', 2, 5, 1],
    ['完成品001', 'ノートパソコンセット', '部品002', '部品002', 10, 2, 4],
    ['完成品001', 'ノートパソコンセット', '部品003', '部品003', 1, 1, 2],
    ['完成品002', 'デスクトップセット', '部品004', '部品004', 5, 1, 1],
    ['完成品002', 'デスクトップセット', '部品005', '部品005', 20, 1, 3],
    ['完成品003', 'マウスセット', '部品006', '部品006', 100, 1, 5],
    ['完成品004', 'キーボードセット', '部品007', '部品007', 50, 2, 2],
    ['完成品004', 'キーボードセット', '部品008', '部品008', 30, 1, 1],
]

wb = Workbook()
ws = wb.active
ws.title = "Parts Master"

# ヘッダー
headers = ['完成品コード', '完成品商品名', '構成部品商品コード（確認用）', '構成部品商品コード', '入数', '箱数', '構成数量']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# データ
for row_idx, row_data in enumerate(master_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 列幅調整
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12

wb.save('/workspaces/-/sample_master.xlsx')
print("✓ sample_master.xlsx を作成しました")

# 照合用ファイル作成
matching_data = [
    ['完成品001'],
    ['完成品002'],
    ['存在しないコード'],
    ['完成品004'],
    ['不正なコード123'],
    ['完成品003'],
]

wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Products"

ws2.cell(row=1, column=1, value="商品コード")
ws2.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
ws2.cell(row=1, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

for idx, data in enumerate(matching_data, 2):
    ws2.cell(row=idx, column=1, value=data[0])

ws2.column_dimensions['A'].width = 20

wb2.save('/workspaces/-/sample_matching.xlsx')
print("✓ sample_matching.xlsx を作成しました")

print("\nサンプルファイルの作成が完了しました！")
